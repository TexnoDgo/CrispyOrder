# Python
import zipfile
import os
from fpdf import FPDF, HTMLMixin
from fpdf import fpdf
import json
import shutil
from openpyxl import Workbook
# Django
from django.shortcuts import render, redirect
from django.views.generic import ListView, CreateView, DetailView, UpdateView, DeleteView
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.mail import send_mail, EmailMessage
from django.conf import settings
from django.utils.translation import ugettext_lazy as _
from django.forms import modelformset_factory
from django.http import HttpResponse, Http404
from django.db.models import Q
# Apps
from chat.models import Message
from chat.forms import MessageCreateForm
from users.models import Profile
# Local
from .handlers import convert_pdf_to_bnp, create_order_pdf
# -------------------------------------------------------OLD MODELS----------------------------------------------------
from .forms import SendOrderForm
# -------------------------------------------------------OLD MODELS----------------------------------------------------

# -------------------------------------------------------NEW MODELS----------------------------------------------------
from .models import CODCity, CODMaterial, CODCategories, CODOrder, CODDetail, CODFile, File
from suggestions.models import CODSuggestion, CODFeedback
from .forms import SingleOrderCreateForm, MultipleOrderCreateForm, AddedOneDetailForm
from .filter import OrdersFilter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


# Отправка заказа другу
@login_required
def send_order_to_friend(request, pk):
    order = CODOrder.objects.get(pk=pk)

    if request.method == 'POST':
        form = SendOrderForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data.get('email', None)
            subject = _('CrispyMachine')
            message = _('An order has been sent to you!')

            data = [[_('Author order:'), str(order.author)],
                    [_('Description:'), str(order.description)],
                    [_('Order created:'), str(order.date_create)],
                    [_('Lead time:'), str(order.lead_time)],
                    [_('Number of items:'), str(order.amount)],
                    [_('Order Budget:'), str(order.proposed_budget)],
                    [_('Order City:'), str(order.city)],
                    [_('Order reference:'), 'http://54.93.42.116:8000/orders/{}'.format(str(order.pk))]
                    ]
            fpdf.set_global("SYSTEM_TTFONTS", os.path.join(os.path.dirname(__file__), 'fonts'))
            ordef_image_view = str(order.image_view)
            pdf_order_url = create_order_pdf(order.image_view.path, data, ordef_image_view)
            print(pdf_order_url)
            email_from = settings.EMAIL_HOST_USER
            recipient_list = [email, ]
            msg = EmailMessage(subject, message, email_from, recipient_list)
            msg.attach_file(pdf_order_url)
            msg.send()
            # message.attach_file = pdf_order_url
            # send_mail(subject, message, email_from, recipient_list)
        return redirect('orders')

    else:
        form = SendOrderForm()
    context = {
        'order': order,
        'form': form,
    }
    return render(request, 'orders/send_order_to_friend.html', context)


# -------------------------------------------------------NEW MODELS----------------------------------------------------
def all_cod_order_view(request):
    # Загрузка моделей
    categories = CODCategories.objects.all()
    city = CODCity.objects.all()

    f = OrdersFilter(request.GET, queryset=CODOrder.objects.all().order_by('-date_create'))

    paginator = Paginator(f.qs, 6)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'categories': categories,
        'city': city,
        'filter': f,
        'page_obj': page_obj,
    }
    return render(request, 'orders/AllOrderPage.html', context)


@login_required
def create_single_order(request):
    categories = CODCategories.objects.all()
    if request.method == 'POST':
        form = SingleOrderCreateForm(request.POST, request.FILES)

        if form.is_valid():
            order = form.save(commit=False)
            order.author = request.user
            form.save()
            pdf_file_name = str(order.pdf_cover)
            png_file_name = '{}{}'.format(pdf_file_name[20:-3], 'png')
            png_full_path = os.path.join(BASE_DIR, 'media/COD_order_image_cover/') + png_file_name
            convert_pdf_to_bnp(order.pdf_cover.path, png_full_path)
            png_path_name = 'COD_order_image_cover/' + png_file_name
            print(png_path_name)
            print(order.pdf_cover)
            order.image_cover = png_path_name
            order.save()

            title = form.cleaned_data.get('title')  # Получение названи заказка из формы
            messages.success(request,
                             # Формирование сообщения со вложенным именем
                             f'You order has been created!Wait for a response! ')
            url = order.pk
            return redirect('views/single_detail/{}'.format(url))
    else:
        form = SingleOrderCreateForm()

    context = {
        'form': form,
        'categories': categories,
    }

    return render(request, 'orders/create_single_order.html', context)


@login_required
def create_multiple_order(request):
    if request.method == 'POST':
        form = MultipleOrderCreateForm(request.POST, request.FILES)

        if form.is_valid():
            order = form.save(commit=False)
            order.author = request.user
            order.group_status = True
            order.status = 'Discussion'
            form.save()
            # Создание обложки заказа
            pdf_file_name = str(order.pdf_cover)
            png_file_name = '{}{}'.format(pdf_file_name[20:-3], 'png')
            png_full_path = os.path.join(BASE_DIR, 'media/COD_order_image_cover/') + png_file_name
            convert_pdf_to_bnp(order.pdf_cover.path, png_full_path)
            png_path_name = 'COD_order_image_cover/' + png_file_name
            order.image_cover = png_path_name
            order.save()

            # Работа с арихивом
            zip_archive = zipfile.ZipFile(order.archive, 'r')
            extract_archive_path = os.path.join(BASE_DIR, 'media/temp/') + str(order.archive)
            zip_archive.extractall(extract_archive_path)
            file_path = os.walk(extract_archive_path)
            folder = []
            data = {}
            for file in file_path:
                folder.append(file)
            file_in_archive = []

            for address, dirs, files in folder:
                for file in files:
                    file_name = str(file)
                    file_path_name = str(address + '/' + file)
                    file_name = file_name.rsplit(".", 1)[0]
                    if file_name not in data:
                        data[file_name] = [file]
                    elif file_name in data:
                        data[file_name].append(file)
                    else:
                        print('Error')
                    file_in_archive.append(file)
            # Создание деталей из файлов архива
            for a in data:
                detail = CODDetail()
                detail.order = order
                detail.Availability_date = detail.Deadline
                detail.name = a
                detail.save()

                for element in data[a]:
                    # Записать файлы из data в заказ
                    #detail_file = File()
                    #detail_file.file = extract_archive_path + '/' + element
                    #detail_file.detail = detail
                    #detail_file.save()
                    # Создание обложки
                    halyard = element[-3:]
                    if halyard == 'PDF':
                        detail_png_file_name = '{}{}'.format(element[:-3], 'png')
                        print('detail_png_file_name: ' + detail_png_file_name)
                        detail_pdf_full_path = extract_archive_path + '/' + element
                        print('detail_pdf_full_path: ' + detail_pdf_full_path)
                        detail_png_full_path = os.path.join(BASE_DIR, 'media/COD_Detail_image_cover/') + detail_png_file_name
                        convert_pdf_to_bnp(detail_pdf_full_path, detail_png_full_path)
                        detail_png_path_name = 'COD_Detail_image_cover/' + detail_png_file_name
                        detail.image_cover = detail_png_path_name
                        detail_file_full_path_name = 'temp/' + str(order.archive) + '/' + element
                        print(detail_file_full_path_name)
                        detail.pdf = detail_file_full_path_name
                        detail.save()
                    elif halyard == 'DXF':
                        detail_file_full_path_name = 'temp/' + str(order.archive) + '/' + element
                        detail.dxf = detail_file_full_path_name
                        detail.save()
                    elif halyard == 'STP' or halyard == 'STEP' or halyard == 'TEP':
                        detail_file_full_path_name = 'temp/' + str(order.archive) + '/' + element
                        detail.step = detail_file_full_path_name
                        detail.save()
                    elif halyard == 'PART' or halyard == 'PRT' or halyard == 'ART':
                        detail_file_full_path_name = 'temp/' + str(order.archive) + '/' + element
                        detail.part = detail_file_full_path_name
                        detail.save()
                    else:
                        print('Error element')
            # Очистка(Удаляет файлы. Убрать очистку)
            #shutil.rmtree(extract_archive_path, ignore_errors=True)

            title = form.cleaned_data.get('title')  # Получение названи заказка из формы
            messages.success(request,
                             # Формирование сообщения со вложенным именем
                             f'You order has been created!Wait for a response! ')
            url = order.pk
            return redirect('view/{}'.format(url))
    else:
        form = MultipleOrderCreateForm()

    context = {
        'form': form,
    }

    return render(request, 'orders/create_multiple_order.html', context)


@login_required
def added_one_detail(request, url):
    added_order = CODOrder.objects.get(pk=url)
    print(added_order.image_cover)
    if request.method == 'POST':
        form = AddedOneDetailForm(request.POST, request.FILES)

        if form.is_valid():
            detail = form.save(commit=False)
            detail.order = CODOrder.objects.get(pk=url)
            detail.Availability_date = detail.Deadline
            added_order.status = 'Discussion'
            detail.save()
            added_order.save()

            title = form.cleaned_data.get('title')  # Получение названи заказка из формы
            messages.success(request,
                             # Формирование сообщения со вложенным именем
                             f'You order has been created!Wait for a response! ')
            return redirect('all_cod_order_view')
    else:
        form = AddedOneDetailForm()

    context = {
        'added_order': added_order,
        'form': form,
    }

    return render(request, 'orders/added_one_detail.html', context)


@login_required
def added_multiple_detail(request, url):
    added_order = CODOrder.objects.get(pk=url)

    DetailFormset = modelformset_factory(CODDetail, fields=('order', 'name', 'amount', 'material', 'whose_material',
                                                             'Note', 'Deadline', 'Availability_date',
                                                            'pdf', 'dxf', 'step', 'part'))
    if request.method == 'POST':
        formset = DetailFormset(request.POST, request.FILES,
                                queryset=CODDetail.objects.filter(order=added_order))
        print('post')
        if formset.is_valid():
            print('valid')
            formset.save()
        return redirect('all_cod_order_view')
    else:
        formset = DetailFormset(queryset=CODDetail.objects.filter(order=CODOrder.objects.get(pk=url)))
        print('else')

    context = {
        'added_order': added_order,
        'formset': formset,
    }
    return render(request, 'orders/added_multiple_detail.html', context)


def order_and_suggestion_view(request, url):
    order = CODOrder.objects.get(pk=url)
    details = CODDetail.objects.filter(order__pk=url)
    single_detail = CODDetail.objects.filter(order__pk=url).first()
    suggestions = CODSuggestion.objects.filter(order=order)
    context = {
        'order': order,
        'details': details,
        'single_detail': single_detail,
        'suggestions': suggestions,
    }
    return render(request, 'orders/order_and_suggestion_view.html', context)


class CODDeleteOrderView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = CODOrder

    success_url = '/'

    def test_func(self):
        order = self.get_object()
        if self.request.user == order.author:
            return True
        return False


class CODOrderUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = CODOrder

    success_url = '/'

    def test_func(self):
        order = self.get_object()
        if self.request.user == order.author:
            return True
        return False

    def get_context_data(self, **kwargs):
        context = super(CODOrderUpdateView, self).get_context_data(**kwargs)
        a = self.object.id
        context['details'] = CODDetail.objects.filter(order=a)
        return context

    fields = ['title', 'description', 'city', 'archive', 'image_cover', 'pdf_cover', 'proposed_budget',
              ]

    def form_valid(self, form):
        form.instance.author = self.request.user
        order = form.save(commit=False)
        order.save()
        return super().form_valid(form)


# ------------------------------------------- Статусы заказов и предложений---------------------------------------
def change_status(request, url):
    suggestion = CODSuggestion.objects.get(pk=url)
    order = CODOrder.objects.get(pk=suggestion.order.pk)

    if suggestion.status == 'Discussion' and order.status == 'Discussion':
        suggestion.status = 'InWork'
        order.status = 'InWork'
    elif suggestion.status == 'InWork' and order.status == 'InWork':
        suggestion.status = 'Done'
        order.status = 'Done'
    else:
        print('Error status')

    suggestion.save()
    order.save()

    other_suggestions = CODSuggestion.objects.filter(order=order)
    for other_suggestion in other_suggestions:
        if other_suggestion != suggestion:
            other_suggestion.status = 'Canceled'
            other_suggestion.save()

    return redirect(request.META['HTTP_REFERER'])
# ------------------------------------------- Статусы заказов и предложений---------------------------------------


# -------------------------------------------------------NEW MODELS----------------------------------------------------
def create_xls_project(request, url):
    order = CODOrder.objects.get(pk=url)
    details = CODDetail.objects.filter(order=order)

    # created xls document
    wb = Workbook()
    # grab the active worksheet
    ws = wb.active
    # ORDER CAP
    ws.merge_cells('A1:H1')
    ws['A1'] = 'ORDER'
    ws['A2'] = 'id'
    ws['B2'] = 'author'
    ws['C2'] = 'date create'
    ws['D2'] = 'title'
    ws['E2'] = 'description'
    ws['F2'] = 'city'
    ws['G2'] = 'budget'
    ws['H2'] = 'status'
    # DATA
    ws['A3'] = order.id
    ws['B3'] = order.author.username
    ws['C3'] = order.date_create
    ws['D3'] = order.title
    ws['E3'] = order.description
    ws['F3'] = str(order.city)
    ws['G3'] = order.proposed_budget
    ws['H3'] = order.status

    # DETAILS CAP
    ws.merge_cells('A5:H5')
    ws['A5'] = 'DETAILS'
    ws['A6'] = 'name'
    ws['B6'] = 'amount'
    ws['C6'] = 'material'
    ws['D6'] = 'whose material'
    ws['E6'] = 'note'
    ws['G6'] = 'deadline'
    ws['H6'] = 'availability date'
    ws['I6'] = 'PDF file'
    ws['J6'] = 'DXF file'
    ws['K6'] = 'STEP file'
    ws['L6'] = 'PART file'
    row = 7
    for detail in details:
        files = File.objects.filter(detail=detail)
        pdf = 'None'
        dxf = 'None'
        step = 'None'
        part = 'None'
        for file in files:
            file_name = str(file)
            if file_name[-3:] == 'PDF':
                pdf = str(file.file)
            elif file_name[-3:] == 'DXF':
                dxf = str(file.file)
            elif file_name[-4:] == 'STEP' or file_name[-3:] == 'STP':
                step = str(file.file)
            elif file_name[-4:] == 'PART' or file_name[-3:] == 'PRT':
                part = str(file.file)

        cell_a = ws.cell(row=row, column=1, value=detail.name)
        cell_b = ws.cell(row=row, column=2, value=detail.amount)
        cell_c = ws.cell(row=row, column=3, value=str(detail.material))
        cell_d = ws.cell(row=row, column=4, value=detail.whose_material)
        cell_e = ws.cell(row=row, column=5, value=detail.Note)
        cell_g = ws.cell(row=row, column=7, value=detail.Deadline)
        cell_h = ws.cell(row=row, column=8, value=detail.Availability_date)
        cell_i = ws.cell(row=row, column=9, value=pdf)
        cell_i.hyperlink = str(pdf)
        cell_j = ws.cell(row=row, column=10, value=dxf)
        cell_j.hyperlink = dxf
        cell_k = ws.cell(row=row, column=11, value=step)
        cell_k.hyperlink = step
        cell_l = ws.cell(row=row, column=12, value=part)
        cell_l.hyperlink = part
        row += 1
    table_path = os.path.join(settings.MEDIA_ROOT, 'temp', "{}.xlsx".format(order.title))
    obj = wb.save(table_path)
    print(table_path)
    with open(table_path, 'rb') as ft:
        response = HttpResponse(ft.read(), content_type="application/vnd.ms-excel")
        response['Content-Disposition'] = 'inline; filename=' + os.path.basename(table_path)
        return response


class CODDetailDelete(LoginRequiredMixin, UserPassesTestMixin, DeleteView):

    model = CODDetail

    success_url = '/'

    def test_func(self):
        detail = self.get_object()
        if self.request.user == detail.order.author:
            return True
        return False

    def get_context_data(self, **kwargs):
        context = super(CODDetailDelete, self).get_context_data(**kwargs)
        a = self.object.id
        detail = self.get_object()
        detail_order = detail.order
        context['detail'] = detail
        context['detail_order'] = detail_order
        return context


class CODDetailUpdate(LoginRequiredMixin, UserPassesTestMixin, UpdateView):

    model = CODDetail

    success_url = '/'

    def test_func(self):
        detail = self.get_object()
        if self.request.user == detail.order.author:
            return True
        return False

    def get_context_data(self, **kwargs):
        context = super(CODDetailUpdate, self).get_context_data(**kwargs)
        a = self.object.id
        detail = self.get_object()
        detail_order = detail.order
        context['detail'] = detail
        context['detail_order'] = detail_order
        return context

    fields = ['name', 'amount', 'material', 'whose_material', 'Note', 'Deadline', 'Availability_date',
              'pdf', 'dxf', 'step', 'part', 'order']

    def form_valid(self, form):
        form.instance.author = self.request.user
        order = form.save(commit=False)
        order.save()
        return super().form_valid(form)